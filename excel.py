'''Write a OOP program to interact with Excel file format which have these functions:
    1) Print out the absolute path to the file
    2) Print the list of sheets and current active sheets
    3) Alter an value in a specific (sheet, cell) Thay đổi giá trị trong ( trang tính, ô) cụ thể  
    os.path.join()
    4) Print all address of specific value (sheet, cell)
    5) Count how many values on a specific rows/columns with 2 options: all available values or specific values 
    
    last updated: 28/03/2023 17:05:00 UTC+7
    written by: Nguyen Thi Minh Chau
'''

from openpyxl import load_workbook
import openpyxl 
import os 


class excel_python :
    def __init__(self, filepath):
        #constructor
        self.filepath = filepath
        self.workbook = openpyxl.load_workbook(filepath)

    def print_absolute_path_file(self):
        '''
        this method prints the absolute path
        '''
        return os.path.abspath(self.filepath)
    
    def print_list_of_sheets(self):
        '''
        this method print list of sheets and current active sheets
        output:
            - name of sheets
            - active sheets
        '''
        wb = self.workbook.sheetnames
        active = self.workbook.active
        return [wb,active]  #return list of sheets and active sheets
              
    def alter_value(self, sheet, cell, value):
        '''
        this method alter value in the sheet 
        input: 
            - name of sheet
            - cell
            - value
        output: *none( save file excel)
        '''
        wb = self.workbook[sheet]
        wb[cell] = value
        self.workbook.save(self.filepath)
        
    def print_address_value(self,value):
        '''
        print all address of specific value
        input: value (string )
        output: list(sheet, value)
        '''
        wb = self.workbook.sheetnames
        return self.__sheet_value(wb,value)
        
    def count_values(self, sheet, value, method, row_column = ""):
        '''
        count values on a sheet with 2 options: all available values or specific values
        input: 
            - name of sheet
            - value (string)
            - method (all, row, column)
        output:
            - number of value 
        ''' 
        count = 0 
        if method == 'all':
            if row_column != None:
                row_column = None
            return self.__count_value_all(sheet, count, value)
        elif method == 'row':
            return self.__count_value_row(count, sheet, value, row_column)
        elif method == 'column':
            return self.__count_value_column(count, sheet, value, row_column) 
        
           
    def __sheet_value (self, wb,value):
        list_sheet_cell = []
        for sheet in wb:
            worksheet = self.workbook[sheet]
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value == value:
                        list = (sheet,cell.coordinate)
                        list_sheet_cell.append(list)
        return list_sheet_cell
        
    def __count_value_all(self,sheet, count, value):
        '''
        count value on a sheet with options: all
        '''
        worksheet = self.workbook[sheet]
        for row in worksheet.iter_rows():
            for cell in row: 
                if cell.value == value:
                    count = count + 1
        return  count
        
    def __count_value_row(self, count, sheet, value, row_column):
        '''
        count value on a sheet with option row
        '''
        worksheet = self.workbook[sheet]
        for r in worksheet.iter_rows(max_row = row_column, min_row = row_column):
            for cell in r: 
                if cell.value == value:
                    count = count + 1
        return count 
    
    def __count_value_column(self,count, sheet, value, row_column):
        '''
        count value on a sheet with option column
        '''
        worksheet = self.workbook[sheet]
        row_column = row_column.upper()
        row_column = ord(row_column) - 64
        for c in worksheet.iter_cols( max_col = row_column, min_col= row_column):
            for cell in c: 
                if cell.value == value:
                    count = count + 1
        return count 
        
################################################################
#test 1 - 5
x = excel_python("python_excel.xlsx")
#print(x.print_absolute_path_file())
#print (x.print_list_of_sheets())
#x.alter_value('Sheet1','A7','hihihi')
#print ('address of value: ',x.print_address_value('goodbye!!!'))
print('count value: ',x.count_values('Sheet1','goodbye!!!','all',''))

    
