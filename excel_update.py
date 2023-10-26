#Write a OOP program to interact with Excel file format which have these functions:
from openpyxl import load_workbook
import openpyxl 
import os
import xlrd

class excel_python :
    #constructor, hàm khởi tạo
    def __init__(self):
        self.workbook = openpyxl.load_workbook("python_excel.xlsx")
        ''' Wrong. The init should take the file path from user not by developer'''

    #1) Print out the absolute path to the file
    def print_absolute_path_file(self):
        return os.path.abspath("python_excel.xlsx")
        """ Wrong distance"""
        
    #2) Print the list of sheets and current active sheets
    def print_list_of_sheets(self,sheet):
        print (self.workbook.sheetnames)
        wb = self.workbook.active 
        wb= self.workbook[sheet]
        print (wb)
        """
        1) Return only. A method should not print the result but return them
        2) Wrong active sheets. DO NOT HARD-CODE
        """
        
#3) Alter an value in a specific (sheet, cell) 
    def alter_value(self, sheet, cell, value):
        wb = self.workbook.active 
        wb= self.workbook[sheet]
        wb[cell] = value
        self.workbook.save("python_excel.xlsx")
        """
        This needs a value input by users not by dev
        """

        
#4) Print all address of specific value (sheet, cell)
    def print_address_value(self, sheet, value):
        workbook = openpyxl.load_workbook("python_excel.xlsx")
        worksheet = workbook[sheet]
        component_value = value
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == component_value:
                    print(cell.coordinate)
        """
        The input should be a value and output should be a list of (sheet,cell) address
        Example: input : value == 'hello' => output == [('Sheet1', 'A12'), ('Sheet2', 'A23'), ('Sheet2', 'B12')]
        """
                          
#5) Count how many values on a specific rows/columns with 2 options: all available values or specific values
    def count_values(self, sheet, value):
        workbook = openpyxl.load_workbook()
        worksheet = workbook[sheet]  
        component_value = value   
        count = 0
        for row in worksheet.iter_rows():
            for cell in row: 
                print ('cell',cell)   
                print ('row',row)     
          
        """
        Determine the quest "A Specific rows/columns". Then in this case, method shoud take a rows/column data
        Next, all available value/specific values. We can use default parameter to guide the user on how to use.
        Finally, the output of this method is depended on dev, you can export a dictionary which have a count of values or a nested list
        """        
################################################################
x = excel_python()

x. count_values('Sheet1','goodbye!!!')