#Write a OOP program to interact with Excel file format which have these functions:
from openpyxl import load_workbook
import openpyxl 
import os 
import xlrd

class excel_python :
    #constructor, hàm khởi tạo
    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = openpyxl.load_workbook(filepath)

    #1) Print out the absolute path to the file
    def print_absolute_path_file(self):
        return os.path.abspath(self.filepath)
    
    #2) Print the list of sheets and current active sheets
    def print_list_of_sheets(self):
        workbook = openpyxl.load_workbook(self.filepath)
        wb = workbook.sheetnames
        active = workbook.active
        return [wb,active]
             
    #3) Alter an value in a specific (sheet, cell) 
    def alter_value(self, sheet, cell, value):
        wb = self.workbook.active 
        wb= self.workbook[sheet]
        wb[cell] = value
        
    #4) Print all address of specific value (sheet, cell)
    def print_address_value(self,value):
        workbook = openpyxl.load_workbook(self.filepath)
        wb = workbook.sheetnames
        list_sheet = []
        list_sheet_cell = []
        for sheet in wb:
            list_sheet.append(sheet)
            for namesheet in list_sheet:
                worksheet = workbook[namesheet]
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value == value:
                            list = (sheet,cell.coordinate)
                            list_sheet_cell.append(list)
                return list_sheet_cell
    
#5) Count how many values on a specific rows/columns with 2 options: all available values or specific values
    def count_values(self, sheet, value, method, row, column):
        workbook = openpyxl.load_workbook("python_excel.xlsx")  
        worksheet = workbook[sheet]  
        count = 0
        if method == 'all':
            for row in worksheet.iter_rows():
                for cell in row: 
                    if cell.value == value:
                        count = count + 1
            return count 
        elif method == 'row':
            for r in worksheet.iter_rows(max_row = row, min_row = row):
                for cell in r: 
                    if cell.value == value:
                        count = count + 1
            return count 
        elif method == 'column':
            column = column.upper()
            column = ord(column) - 64
            for c in worksheet.iter_cols( max_col = column, min_col= column):
                for cell in c: 
                    if cell.value == value:
                        count = count + 1
            return count 
        
################################################################
#test 1 - 4
x = excel_python("python_excel.xlsx")
print(x.print_absolute_path_file())
print (x.print_list_of_sheets())
x.alter_value('Sheet1','A1','hihihi')
print (x.print_address_value('goodbye!!!'))
print(x.count_values('Sheet1','goodbye!!!','row',7,'G'))

    
